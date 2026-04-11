from flask import Flask, request, jsonify, render_template, Response, send_file
import scraper
import cache
import atexit
import json
import uuid
import os
import io
import time
import threading

app = Flask(__name__, template_folder='templates', static_folder='static')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads')
RESULT_DIR = os.path.join(BASE_DIR, 'results')
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(RESULT_DIR, exist_ok=True)

SEARCH_CACHE_TTL = 300      # 5 menit
DETAIL_CACHE_TTL = 3600     # 1 jam

# Bulk job tracking
bulk_jobs = {}


# ============================================================
# TENAGA KERJA ROUTES
# ============================================================

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/search')
def api_search():
    q = request.args.get('q', '').strip()
    search_type = request.args.get('type', 'Nama')

    if not q:
        return jsonify({'error': 'Parameter q (keyword) harus diisi'}), 400
    if search_type not in ('Nama', 'NIK'):
        return jsonify({'error': 'Parameter type harus Nama atau NIK'}), 400

    cache_key = f'search:{search_type}:{q}'
    cached = cache.get(cache_key, SEARCH_CACHE_TTL)
    if cached:
        cached['from_cache'] = True
        return jsonify(cached)

    try:
        result = scraper.search_tenaga_kerja(q, search_type)
        cache.set(cache_key, result)
        result['from_cache'] = False
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/detail')
def api_detail():
    path = request.args.get('path', '').strip()
    if not path:
        return jsonify({'error': 'Parameter path harus diisi'}), 400

    cache_key = f'detail:{path}'
    cached = cache.get(cache_key, DETAIL_CACHE_TTL)
    if cached:
        return jsonify(cached)

    try:
        result = scraper.get_detail(path)
        cache.set(cache_key, result)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================
# BADAN USAHA ROUTES
# ============================================================

@app.route('/api/search-bu')
def api_search_bu():
    q = request.args.get('q', '').strip()
    search_type = request.args.get('type', 'Nama')

    if not q:
        return jsonify({'error': 'Parameter q (keyword) harus diisi'}), 400
    if search_type not in ('Nama', 'NPWP'):
        return jsonify({'error': 'Parameter type harus Nama atau NPWP'}), 400

    cache_key = f'search_bu:{search_type}:{q}'
    cached = cache.get(cache_key, SEARCH_CACHE_TTL)
    if cached:
        cached['from_cache'] = True
        return jsonify(cached)

    try:
        result = scraper.search_badan_usaha(q, search_type)
        cache.set(cache_key, result)
        result['from_cache'] = False
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/detail-bu')
def api_detail_bu():
    path = request.args.get('path', '').strip()
    if not path:
        return jsonify({'error': 'Parameter path harus diisi'}), 400

    cache_key = f'detail_bu:{path}'
    cached = cache.get(cache_key, DETAIL_CACHE_TTL)
    if cached:
        return jsonify(cached)

    try:
        result = scraper.get_detail_badan_usaha(path)
        cache.set(cache_key, result)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================
# BULK CHECK
# ============================================================

def process_bulk(job_id, names, search_type):
    """Background bulk check worker."""
    try:
        bulk_jobs[job_id]['status'] = 'processing'
        total = len(names)
        results = []

        for i, name in enumerate(names):
            name = name.strip()
            if not name:
                continue

            entry = {'input': name, 'found': False, 'profil': {}, 'sertifikat': []}

            try:
                search_data = scraper.search_tenaga_kerja(name, search_type)
                if search_data['results']:
                    best = search_data['results'][0]
                    entry['found'] = True
                    entry['nama'] = best.get('nama', '')
                    entry['nik'] = best.get('nik', '')
                    entry['lokasi'] = best.get('lokasi', '')

                    try:
                        detail = scraper.get_detail(best['detail_url'])
                        entry['profil'] = detail.get('profil', {})
                        entry['sertifikat'] = detail.get('sertifikat', [])
                    except Exception:
                        pass

                    time.sleep(0.5)
                else:
                    entry['error'] = 'Tidak ditemukan'
            except Exception as e:
                entry['error'] = str(e)

            results.append(entry)
            bulk_jobs[job_id]['current'] = i + 1
            bulk_jobs[job_id]['total'] = total
            bulk_jobs[job_id]['logs'].append({
                'index': i + 1, 'total': total, 'nama': name,
                'status': 'Ditemukan' if entry['found'] else entry.get('error', 'Tidak ditemukan'),
            })
            if len(bulk_jobs[job_id]['logs']) > 500:
                bulk_jobs[job_id]['logs'] = bulk_jobs[job_id]['logs'][-500:]
            time.sleep(0.3)

        bulk_jobs[job_id]['results'] = results
        bulk_jobs[job_id]['status'] = 'done'
        _generate_bulk_excel(job_id, results)

    except Exception as e:
        bulk_jobs[job_id]['status'] = 'error'
        bulk_jobs[job_id]['error'] = str(e)


def _generate_bulk_excel(job_id, results):
    """Generate Excel output from bulk check results."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hasil Cek SKA"

    headers = ["No", "Input Pencarian", "Status", "Nama (LPJK)", "NIK", "Lokasi",
               "Pendidikan", "Total Sertifikat", "Sertifikat Aktif", "Detail Sertifikat"]

    hdr_fill = PatternFill(start_color="4c6ef5", end_color="4c6ef5", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for i, r in enumerate(results, 2):
        aktif = sum(1 for s in r.get('sertifikat', []) if s.get('status') == 'Berlaku')
        detail_parts = [f"{s.get('tipe','-')}: {s.get('judul','-')} ({s.get('status','-')})"
                        for s in r.get('sertifikat', [])]
        row_data = [
            i - 1, r.get('input', ''),
            'Ditemukan' if r.get('found') else 'Tidak Ditemukan',
            r.get('profil', {}).get('nama', r.get('nama', '')),
            r.get('profil', {}).get('nik', r.get('nik', '')),
            r.get('lokasi', ''),
            r.get('profil', {}).get('pendidikan', ''),
            r.get('profil', {}).get('total_sertifikat', ''),
            aktif,
            '; '.join(detail_parts) if detail_parts else '-',
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            cell.font = Font(size=10, name="Arial")
            if col == 3:
                cell.fill = green_fill if r.get('found') else red_fill

    widths = [5, 25, 16, 30, 20, 25, 15, 15, 15, 60]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w
    ws.freeze_panes = "A2"

    output_path = os.path.join(RESULT_DIR, f"{job_id}_hasil_cek_ska.xlsx")
    wb.save(output_path)
    bulk_jobs[job_id]['output_path'] = output_path


@app.route('/api/bulk-upload', methods=['POST'])
def api_bulk_upload():
    search_type = request.form.get('type', 'Nama')

    if 'file' in request.files:
        f = request.files['file']
        if f.filename.lower().endswith('.xlsx'):
            import openpyxl
            tmp = os.path.join(UPLOAD_DIR, f"tmp_{f.filename}")
            f.save(tmp)
            wb = openpyxl.load_workbook(tmp, read_only=True)
            ws = wb.active
            names = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = row[1] if len(row) > 1 else row[0]
                if val:
                    names.append(str(val).strip())
            wb.close()
            os.remove(tmp)
        else:
            return jsonify({'error': 'File harus .xlsx'}), 400
    else:
        text = request.form.get('names', '')
        names = [n.strip() for n in text.split('\n') if n.strip()]

    if not names:
        return jsonify({'error': 'Tidak ada data untuk dicek'}), 400

    job_id = str(uuid.uuid4())[:8]
    bulk_jobs[job_id] = {
        'status': 'queued', 'current': 0, 'total': len(names),
        'logs': [], 'results': [], 'error': None, 'output_path': None,
    }

    thread = threading.Thread(target=process_bulk, args=(job_id, names, search_type), daemon=True)
    thread.start()
    return jsonify({'job_id': job_id, 'total': len(names)})


@app.route('/api/bulk-progress/<job_id>')
def api_bulk_progress(job_id):
    def generate():
        retries = 0
        while job_id not in bulk_jobs and retries < 10:
            time.sleep(0.5)
            retries += 1
        if job_id not in bulk_jobs:
            yield f"data: {json.dumps({'error': 'Job tidak ditemukan'})}\n\n"
            return
        last_sent = 0
        while True:
            job = bulk_jobs.get(job_id)
            if not job:
                break
            logs_to_send = job['logs'][last_sent:]
            last_sent = len(job['logs'])
            payload = {'status': job['status'], 'current': job['current'],
                       'total': job['total'], 'logs': logs_to_send}
            if job['status'] == 'done':
                found = sum(1 for r in job['results'] if r.get('found'))
                payload['summary'] = {'total': len(job['results']),
                                      'found': found, 'not_found': len(job['results']) - found}
                yield f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"
                break
            elif job['status'] == 'error':
                payload['error'] = job['error']
                yield f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"
                break
            yield f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"
            time.sleep(1)

    return Response(generate(), mimetype='text/event-stream',
                    headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})


@app.route('/api/bulk-download/<job_id>')
def api_bulk_download(job_id):
    job = bulk_jobs.get(job_id)
    if job and job.get('output_path') and os.path.exists(job['output_path']):
        return send_file(job['output_path'], as_attachment=True, download_name='Hasil_Cek_SKA.xlsx')
    return jsonify({'error': 'File tidak ditemukan'}), 404


# ============================================================
# EXPORT
# ============================================================

@app.route('/api/export', methods=['POST'])
def api_export():
    """Export search results or detail to Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side

    data = request.json
    if not data:
        return jsonify({'error': 'No data'}), 400

    wb = openpyxl.Workbook()
    ws = wb.active
    hdr_fill = PatternFill(start_color="4c6ef5", end_color="4c6ef5", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    if data.get('type') == 'search':
        ws.title = "Hasil Pencarian"
        headers = ["No", "Nama", "NIK", "Lokasi"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill, c.font, c.border = hdr_fill, hdr_font, thin_border
        for i, r in enumerate(data.get('results', []), 2):
            for col, val in enumerate([i-1, r.get('nama',''), r.get('nik',''), r.get('lokasi','')], 1):
                ws.cell(row=i, column=col, value=val).border = thin_border

    elif data.get('type') == 'detail':
        ws.title = "Detail Sertifikat"
        profil = data.get('profil', {})
        ws.cell(row=1, column=1, value="Nama").font = Font(bold=True)
        ws.cell(row=1, column=2, value=profil.get('nama', ''))
        ws.cell(row=2, column=1, value="NIK").font = Font(bold=True)
        ws.cell(row=2, column=2, value=profil.get('nik', ''))
        ws.cell(row=3, column=1, value="Pendidikan").font = Font(bold=True)
        ws.cell(row=3, column=2, value=profil.get('pendidikan', ''))

        headers = ["No", "Tipe", "Judul", "Status", "Jenjang", "Sub Bidang",
                   "Klasifikasi", "Kualifikasi", "Berlaku Hingga", "No. Registrasi"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=5, column=col, value=h)
            c.fill, c.font, c.border = hdr_fill, hdr_font, thin_border
        for i, s in enumerate(data.get('sertifikat', []), 6):
            vals = [i-5, s.get('tipe',''), s.get('judul',''), s.get('status',''),
                    s.get('jenjang',''), s.get('sub_bidang',''), s.get('klasifikasi',''),
                    s.get('kualifikasi',''), s.get('berlaku_hingga',''), s.get('nomor_registrasi','')]
            for col, val in enumerate(vals, 1):
                ws.cell(row=i, column=col, value=val).border = thin_border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = data.get('filename', 'export') + '.xlsx'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================================
# CLEANUP
# ============================================================

@atexit.register
def cleanup():
    scraper.close()


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 3000))
    print(f'=== SKA Scanner by Bashid Effendi ===')
    print(f'Server berjalan di http://localhost:{port}')
    app.run(host='0.0.0.0', port=port, debug=False)
